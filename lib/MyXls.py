import openpyxl
from .Validate import Validate

class MyXls(Validate):
	_bodyXLS = []
	_hBook = ''
	_hSheet = ''
	_errors = []
	
	def __init__(self, fileName, sheetName):
		self._hBook = openpyxl.load_workbook(fileName)
		self._hSheet = self._hBook.get_sheet_by_name(sheetName)		
		self._generate()
	
	def _generate(self):
		for i in range(1, self._hSheet.max_row):
			rows = []
			field = ''
			for j in range(1, self._hSheet.max_column):
				fieldNew = self._hSheet.cell(row=i,column=j).value
				fieldNew = self.verify(fieldNew)
			
				if (len(self.errorValidate()) > 0):
					self._errors.append(self.errorValidate())

				rows.append(fieldNew)
			self._bodyXLS.append(rows)

	def results(self,flag='body'):
		if flag == 'header':
			return self._bodyXLS.pop(0)
		return self._bodyXLS

	def counts(self):
		return len(self._bodyXLS)

	def errors(self):
		return self._errors

	def errorsCount(self):
		return len(self._errors)